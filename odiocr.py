# coding: utf-8
# GNU GPL v3
# Icone de l'application en partie extraite du site https://www.flaticon.com et https://icon-icons.com/
# Les bibliothèques importées
###
# Commande compilation pyinstaller
# pyi-makespec odiocr.py    création de la configuration de la compilation
#
# pyinstaller.exe main.spec --noconfirm     compilation avec fenetre de debuggage
#
# pyinstaller.exe odiocr.spec --onefile --noconfirm     compilation de production
###

import textwrap
from fenetre_maisonretraite import FenetreDossierMaisonRetraite
from fenetre_donnee import FenetreDonnee
from fenetre_anamnese import FenetreAnamnese
import keyboard
from pynput.keyboard import *
import time
import pyautogui
import os
import sys
import pygame
import openpyxl
from openpyxl.styles import Font, Color
from datetime import datetime
import shutil
from pathlib import Path
from fonctiondekev import pdf_merge
from fonctiondekev import loss_noah_extractor
from fonctiondekev import excel_triage
from fonctiondekev import showMessage
from fonctiondekev import excel_close_test
from tkinter import messagebox
import pickle

# Variable globale pour gérer les étapes dans l'ordre
global step_one, step_two, step_three, numero_patient, nom_maison_retraite, nom_patient, prenom_patient, nom_accompagnant, prenom_accompagnant, telephone_accompagnant, mail_accompagnant, app3
global chemin_calisto, dossier_sauvegarde, la_date_jour_save, choix_mode
# Variable globale pour l'enregistrement dans la base de donnée
global nom_patients, prenom_patients, nom_accompagnants, prenom_accompagnants, telephone_accompagnants, \
    mail_accompagnants, anamnese_db1, anamnese_db2, anamnese_db3, anamnese_db4, anamnese_db5, anamnese_db6, \
    anamnese_db7, anamnese_db8, anamnese_db9, anamnese_db10, anamnese_db11, empreinte_OD, empreinte_OG, oui_color, \
    impossible_color, defaut_color, liste_enregistre_tmp, synthese_enregistre_tmp, reglagecommentaire

nom_patients = [""] * 50
prenom_patients = [""] * 50
nom_accompagnants = [""] * 50
prenom_accompagnants = [""] * 50
telephone_accompagnants = [""] * 50
mail_accompagnants = [""] * 50
anamnese_db1 = [""] * 50
anamnese_db2 = [""] * 50
anamnese_db3 = [""] * 50
anamnese_db4 = [""] * 50
anamnese_db5 = [""] * 50
anamnese_db6 = [""] * 50
anamnese_db7 = [""] * 50
anamnese_db8 = [""] * 50
anamnese_db9 = [""] * 50
anamnese_db10 = [""] * 50
anamnese_db11 = [""] * 50

if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
    print('running in a PyInstaller bundle')
else:
    print('running in a normal Python process')

# Configuration police pour la synthèse
taillepolice = 8
oui_color = Font(color="296c10", size=taillepolice)
impossible_color = Font(color="6c1610", size=taillepolice)
defaut_color = Font(size=taillepolice)

# Implémentation d'un mode test qui marche sur d'autre machine n'ayant ni l'arborescence ni les fichiers prérequis
# Normal où test
choix_mode = ""


def main():
    global choix_mode
    # Use a breakpoint in the code line below to debug your script.
    for arg in sys.argv[1:]:
        choix_mode = arg[0:4]


if __name__ == '__main__':
    main()

if choix_mode == "test":
    try:
        os.mkdir(Path("mode_test", "DocumentCalisto"))
        os.mkdir(Path("mode_test", "Depistage"))
    except:
        pass

####
# Path à modifier en fonction de votre configuration
if choix_mode == "test":
    chemin_calisto = Path(Path(__file__).parent.absolute(), "mode_test", "DocumentCalisto")
    chemin_liste_referent = Path(Path(__file__).parent.absolute(), "mode_test", "fichier_test", "Liste Referents.xlsx")
    dossier_sauvegarde = Path(Path(__file__).parent.absolute(), "mode_test", "Depistages")
else:
    chemin_calisto = Path(Path.home(), "Documents", "DocumentCalisto")
    chemin_liste_referent = Path(Path(__file__).parent.absolute(), "mode_test", "fichier_test", "Liste Referents.xlsx")
    chemin_synthese_depistage = Path(Path(__file__).parent.absolute(), "mode_test", "fichier_test",
                                     "synthese_depistage.xlsx")
    dossier_sauvegarde = Path(Path.home(), "Documents", "DocOdipro", "Depistages")
####

pickle.dump(chemin_calisto, open("files/chemin_calisto.dat", "wb"))
pickle.dump(dossier_sauvegarde, open("files/dossier_sauvegarde.dat", "wb"))

# Paramètre fonctionnement
duree = 0.1
step_one = 0
step_two = 0
step_three = 0
numero_patient = 0
#Réglage nombre de caractère ligne commentaire
reglagecommentaire = 110


# Extraction du nom de dossier de la MR GUI pour vérifier l'état du dossier de l'audiomètre il doit contenir
# uniquement le dossier vide avec nom de la maison de retraite
def fermetureapp1():
    if messagebox.askokcancel("Quitter", "Voulez-vous fermer le programme?"):
        app1.destroy()
        sys.exit()

app1 = FenetreDossierMaisonRetraite(chemin_calisto, dossier_sauvegarde)
app1.protocol("WM_DELETE_WINDOW", fermetureapp1)
app1.mainloop()
chemin_calisto = app1.chemin_calisto
dossier_sauvegarde = app1.dossier_sauvegarde

# Tests d'erreur pour vérifier les pré-requis
files = os.listdir(chemin_calisto)
nb_de_fichier = len(files)
files_str = str(files)
if app1.mode_recuperation == 0:
    if nb_de_fichier > 1:
        print("Le dossier n'est pas conforme, trop de fichier(s)/dossier(s)")
        showMessage("Le dossier n'est pas conforme, trop de fichier(s)/dossier(s), fermeture automatique", "warning")
        nom_maison_retraite = ""
        sys.exit()

    elif nb_de_fichier == 0:
        print("Il n'y a pas de dossier avec le nom de la maison de retraite")
        showMessage("Il n'y a pas de dossier avec le nom de la maison de retraite, fermeture automatique", "warning")
        nom_maison_retraite = ""
        sys.exit()

    else:
        # Extraction du nom de dossier
        nom_maison_retraite = app1.nom_maison_retraite

        # On rempli la maison de retraite et la date du jour dans l'excel des référents
        la_date_jour_save = datetime.today().strftime('%d-%m-%Y')
        liste_enregistre_tmp = Path(chemin_calisto, nom_maison_retraite,
                                    "ListeRef-" + nom_maison_retraite + "_temp.xlsx")
        wb_liste_referent = openpyxl.load_workbook(chemin_liste_referent)
        ws_liste_referent = wb_liste_referent['Feuil1']
        sheet_liste_referent = wb_liste_referent.active

        # Adjonction fonction filtrage
        sheet_liste_referent.auto_filter.ref = "B3:I3"
        ws_liste_referent.cell(row=2, column=4).value = app1.nom_maison_retraite
        ws_liste_referent.cell(row=2, column=7).value = la_date_jour_save

        # On rempli la maison de retraite et la date du jour dans la synthese du dépistage
        synthese_enregistre_tmp = Path(chemin_calisto, nom_maison_retraite,
                                       "Synthese-" + nom_maison_retraite + "_temp.xlsx")
        wb_synthese_depistage = openpyxl.load_workbook(chemin_synthese_depistage)
        ws_synthese_depistage = wb_synthese_depistage['Feuil1']
        sheet_synthese_depistage = wb_synthese_depistage.active

        # Adjonction fonction filtrage
        sheet_synthese_depistage.auto_filter.ref = "B5:E5"
        ws_synthese_depistage.cell(row=3, column=3).value = app1.nom_maison_retraite
        ws_synthese_depistage.cell(row=1, column=4).value = la_date_jour_save


# Récupération des informations dans les fichiers excel triage et synthèse si l'option est sélectionnée
elif app1.mode_recuperation == 1:
    print("Récupération patient")

    # On charge le nom de la maison de retraite
    nom_maison_retraite = app1.nom_maison_retraite
    la_date_jour_save = datetime.today().strftime('%d-%m-%Y')

    # On charge la liste des référents de la session précédente
    liste_enregistre_tmp = Path(chemin_calisto, nom_maison_retraite,
                                "ListeRef-" + nom_maison_retraite + "_temp.xlsx")
    wb_liste_referent = openpyxl.load_workbook(liste_enregistre_tmp)
    ws_liste_referent = wb_liste_referent['Feuil1']
    sheet_liste_referent = wb_liste_referent.active

    # On charge la synthèse de la session précédente
    synthese_enregistre_tmp = Path(chemin_calisto, nom_maison_retraite,
                                   "Synthese-" + nom_maison_retraite + "_temp.xlsx")
    wb_synthese_depistage = openpyxl.load_workbook(synthese_enregistre_tmp)
    ws_synthese_depistage = wb_synthese_depistage['Feuil1']
    sheet_synthese_depistage = wb_synthese_depistage.active

    # On analyse la liste pour déterminer le nombre de patients à charger
    for row in sheet_liste_referent.iter_rows(min_row=4, max_row=204, min_col=2, max_col=2):
        for cell in row:
            if not cell.value:
                break
            else:
                numero_patient += 1
                print(cell.value)
    patient_recupere = str(numero_patient)
    showMessage(patient_recupere + " patients récupérés")




# Fonction principale qui détecte les touches du clavier
def press_on(key):
    global step_one, step_two, step_three, numero_patient, nom_maison_retraite, nom_patient, prenom_patient, \
        nom_accompagnant, prenom_accompagnant, telephone_accompagnant, mail_accompagnant, app3, empreinte_OD, \
        empreinte_OG, oui_color, impossible_color, defaut_color, wb_liste_referent, wb_synthese_depistage, \
        ws_liste_referent, ws_synthese_depistage, liste_enregistre_tmp, synthese_enregistre_tmp, analyse_perte

    if key == Key.f9:
        if step_one == 0 and step_two == 0 and step_three == 0:
            # On ouvre la fenetre permettant d'entrer les données patient et accompagnant
            app2 = FenetreDonnee()
            app2.mainloop()
            nom_patient = app2.nom_patient
            prenom_patient = app2.prenom_patient
            nom_accompagnant = app2.nom_accompagnant
            prenom_accompagnant = app2.prenom_accompagnant
            telephone_accompagnant = app2.telephone_accompagnant
            mail_accompagnant = app2.mail_accompagnant

            # On enregistre les informations pour la base de donnée
            nom_patients[numero_patient] = app2.nom_patient
            prenom_patients[numero_patient] = app2.prenom_patient
            nom_accompagnants[numero_patient] = app2.nom_accompagnant
            prenom_accompagnants[numero_patient] = app2.prenom_accompagnant
            telephone_accompagnants[numero_patient] = app2.telephone_accompagnant
            mail_accompagnants[numero_patient] = app2.mail_accompagnant

            if choix_mode == "test":
                pass
            else:
                if app2.comportement == 1:
                    # Première étape on récupère nom prénom et on allume Calisto
                    # Clic fichier
                    pyautogui.click(42, 34)
                    time.sleep(duree)
                    # Clic Ajouter nouveau patient
                    pyautogui.click(118, 99)
                    time.sleep(duree)
                    # Clic Nom patient + remplissage
                    pyautogui.click(750, 232)
                    time.sleep(duree)
                    keyboard.write(nom_patient.upper())
                    time.sleep(duree)
                    # CLic Prénom patient + remplissage
                    pyautogui.click(1149, 232)
                    time.sleep(duree)
                    keyboard.write(prenom_patient.title())
                    time.sleep(duree)
                    # Selection sexe patient
                    if app2.sexe_patient == "Homme":
                        pyautogui.click(1153, 267)
                    else:
                        pyautogui.click(1239, 268)
                    time.sleep(duree)
                    # On enregistre la fiche patient
                    pyautogui.click(1111, 874)
                    time.sleep(duree)
                else:
                    pass

                if app2.comportement == 1 or app2.comportement == 2:
                    # On clique sur la barre de recherche
                    pyautogui.click(22, 171)
                    time.sleep(duree)
                    # On colle le prénom et le nom du patient
                    prenom_nom_patient = prenom_patient.title() + "  " + nom_patient.upper()
                    keyboard.write(prenom_nom_patient)
                    print(prenom_nom_patient)
                    time.sleep(1.5)
                    # Selection fiche
                    pyautogui.click(52, 340)
                    time.sleep(0.6)
                    pyautogui.click(118, 337)
                    time.sleep(0.6)
                    # On ouvre le programme Calisto
                    pyautogui.click(103, 74)
            # On valide la première étape
            step_one = 1

    if key == Key.f10:
        if step_one == 1 and step_two == 0 and step_three == 0:
            # On force l'audiométrie tonale la lecture de l'audiogramme tonal
            pyautogui.click(25, 92)
            time.sleep(0.1)
            # on récupère les pertes moyenne avec pyscreenshot et pytesseract
            print("Analyse de l'audiograme")
            analyse_perte = loss_noah_extractor()
            print("Analyse de l'audiograme terminée")
            # Ouverture de la fenetre graphique
            app3 = FenetreAnamnese(analyse_perte)
            app3.mainloop()
            # Récupération de la modification éventuelle de l'analyse
            analyse_perte = list(analyse_perte)
            analyse_perte[2] = app3.analyse_perte[2]
            if choix_mode == "test":
                pass
            else:
                # On force l'audiométrie tonale pour ouvrir la note
                pyautogui.click(25, 92)
                time.sleep(0.1)
                # Ouverture note
                pyautogui.click(157, 587)
                # Placement correct de la fenêtre
                pyautogui.moveTo(941, 208, 0.1)
                pyautogui.dragTo(317, 411, 0.2, button='left')
                pyautogui.click(8, 463)

                # On transforme les résultats obtenus avec le GUI en texte à copier dans le calisto pour le rapport
                # On integre le nom de la maison de retraite au debut du texte
                text_a_copier = "Etablissement : " + app1.nom_maison_retraite + "\n" + "\n"

                if app3.text_reponse[0] == "Refuse le dépistage":
                    text_a_copier = text_a_copier + "   Refus d'effectuer le dépistage" + "\n"
                    # Ligne otoscopie
                    if app3.text_reponse[3] == "" and app3.text_reponse[4] == "":
                        pass
                    elif app3.text_reponse[3] == "":
                        text_a_copier = text_a_copier + "   " + app3.les_texts[2] + " " + app3.text_reponse[4] + "\n"
                    elif app3.text_reponse[4] == "":
                        text_a_copier = text_a_copier + "   " + app3.les_texts[2] + " " + app3.text_reponse[3] + "\n"
                    else:
                        text_a_copier = text_a_copier + "   " + app3.les_texts[2] + " " + app3.text_reponse[3] + ",  " + \
                                        app3.text_reponse[4] + "\n"
                    # Ligne remarque
                    if app3.text_reponse[5] == "":
                        text_a_copier = text_a_copier + "   " + app3.les_texts[3] + " Aucune" + "\n"
                    else:
                        text_a_copier = text_a_copier + "   " + app3.les_texts[3] + " " + app3.text_reponse[5] + "\n"

                elif app3.text_reponse[0] == "Dépistage impossible":
                    text_a_copier = text_a_copier + "   Impossible d'éffectuer le dépistage" + "\n"
                    # Ligne otoscopie
                    if app3.text_reponse[3] == "" and app3.text_reponse[4] == "":
                        pass
                    elif app3.text_reponse[3] == "":
                        text_a_copier = text_a_copier + "   " + app3.les_texts[2] + " " + app3.text_reponse[4] + "\n"
                    elif app3.text_reponse[4] == "":
                        text_a_copier = text_a_copier + "   " + app3.les_texts[2] + " " + app3.text_reponse[3] + "\n"
                    else:
                        text_a_copier = text_a_copier + "   " + app3.les_texts[2] + " " + app3.text_reponse[3] + ",  " + \
                                        app3.text_reponse[4] + "\n"
                    # Ligne remarque
                    if app3.text_reponse[5] == "":
                        text_a_copier = text_a_copier + "   " + app3.les_texts[3] + " Aucune" + "\n"
                    else:
                        text_a_copier = text_a_copier + "   " + app3.les_texts[3] + " " + app3.text_reponse[5] + "\n"

                else:
                    if app3.text_reponse[0] == "Controle d'appareillage":
                        text_a_copier = text_a_copier + "   Controle d'appareillage auditif" + "\n"

                    # Ligne cognition
                    if app3.text_reponse[1] == "":
                        pass
                    else:
                        text_a_copier = text_a_copier + "   " + app3.les_texts[0] + "  " + app3.text_reponse[1] + "\n"

                    # Ligne déjà appareillé
                    if app3.text_reponse[2] == "":
                        pass
                    else:
                        text_a_copier = text_a_copier + "   " + app3.les_texts[1] + " " + app3.text_reponse[2] + "\n"

                    # Ligne otoscopie
                    if app3.text_reponse[3] == "" and app3.text_reponse[4] == "":
                        pass
                    elif app3.text_reponse[3] == "":
                        text_a_copier = text_a_copier + "   " + app3.les_texts[2] + " " + app3.text_reponse[4] + "\n"
                    elif app3.text_reponse[4] == "":
                        text_a_copier = text_a_copier + "   " + app3.les_texts[2] + " " + app3.text_reponse[3] + "\n"
                    else:
                        text_a_copier = text_a_copier + "   " + app3.les_texts[2] + " " + app3.text_reponse[3] + ",  " + \
                                        app3.text_reponse[4] + "\n"

                    # Décision empreinte gestion dans le cas où il n'y aurait pas d'appareil auditif du tout
                    empreinte_OG = app3.id_reponse[9]
                    empreinte_OD = app3.id_reponse[10]
                    if empreinte_OG == "":
                        empreinte_OG = 1
                    if empreinte_OD == "":
                        empreinte_OD = 1
                    # Ligne prise d'empreinte

                    if empreinte_OG < 6 and empreinte_OD < 6:
                        pass
                    elif empreinte_OG < 6 and empreinte_OD > 5:
                        text_a_copier = text_a_copier + "   Empreinte : OD" + "\n"
                    elif empreinte_OG > 5 and empreinte_OD < 6:
                        text_a_copier = text_a_copier + "   Empreinte : OG" + "\n"
                    elif empreinte_OG > 5 and empreinte_OD > 5:
                        text_a_copier = text_a_copier + "   Empreintes : OG + OD" + "\n"

                    # Ligne remarque
                    if app3.text_reponse[5] == "":
                        text_a_copier = text_a_copier + "   " + app3.les_texts[3] + " Aucune" + "\n"
                    else:
                        text_a_copier = text_a_copier + "   " + app3.les_texts[3] + " " + app3.text_reponse[5] + "\n"
                    # Ligne analyse
                    text_a_copier = text_a_copier + "  Analyse de la perte : " + analyse_perte[2] + "\n"
                    # Ligne avis
                    if app3.text_reponse[6] == "":
                        text_a_copier = text_a_copier + "   " + app3.les_texts[
                            4] + " Impossible de receuillir une réponse" + "\n"
                    elif app3.text_reponse[0] == "Refuse le dépistage" or app3.text_reponse[0] == "Dépistage impossible":
                        pass
                    elif app3.text_reponse[8] == "OD non appareillable" and app3.text_reponse[7] == "OG non appareillable":
                        pass
                    else:
                        text_a_copier = text_a_copier + "   " + app3.les_texts[4] + " " + app3.text_reponse[6] + "\n"

                    # Ligne conseil d'appareillage
                    if app3.text_reponse[7] == "" and app3.text_reponse[8] == "":
                        pass
                    elif app3.text_reponse[7] == "":
                        text_a_copier = text_a_copier + "   " + app3.les_texts[5] + " " + app3.text_reponse[8] + "\n"
                    elif app3.text_reponse[8] == "":
                        text_a_copier = text_a_copier + "   " + app3.les_texts[5] + " " + app3.text_reponse[7] + "\n"
                    else:
                        text_a_copier = text_a_copier + "   " + app3.les_texts[5] + " " + app3.text_reponse[7] + ",  " + \
                                        app3.text_reponse[8] + "\n"

                    # Ligne choix dome embout
                    if app3.text_reponse[9] == "" and app3.text_reponse[10] == "":
                        pass
                    elif app3.text_reponse[9] == "":
                        text_a_copier = text_a_copier + "   " + app3.les_texts[6] + " " + app3.text_reponse[10] + "\n"
                    elif app3.text_reponse[10] == "":
                        text_a_copier = text_a_copier + "   " + app3.les_texts[6] + " " + app3.text_reponse[9] + "\n"
                    else:
                        text_a_copier = text_a_copier + "   " + app3.les_texts[6] + " " + app3.text_reponse[9] + ",  " + \
                                        app3.text_reponse[10] + "\n"

                print(text_a_copier)
                pyautogui.click(5, 461)
                # pyautogui.dragTo(650, 1110, 0.2, button='left')
                keyboard.write(text_a_copier)
                # Enregistrement des informations de l'anamnèse pour la base de donnée

            anamnese_db1[numero_patient] = app3.id_reponse[0]
            anamnese_db2[numero_patient] = app3.id_reponse[1]
            anamnese_db3[numero_patient] = app3.text_reponse[2]
            anamnese_db4[numero_patient] = app3.id_reponse[4]
            anamnese_db5[numero_patient] = app3.id_reponse[3]
            anamnese_db6[numero_patient] = app3.text_reponse[5]
            anamnese_db7[numero_patient] = app3.text_reponse[6]
            anamnese_db8[numero_patient] = app3.id_reponse[10]
            anamnese_db9[numero_patient] = app3.id_reponse[9]
            anamnese_db10[numero_patient] = app3.id_reponse[8]
            anamnese_db11[numero_patient] = app3.id_reponse[7]
            step_two = 1
        else:
            print("Il faut compléter la premiere étape où finir la troisième étape")
            showMessage("Il faut compléter la premiere étape où finir la troisième étape", type='warning', timeout=1500)

    if key == Key.f11:
        if step_one == 1 and step_two == 1 and step_three == 0:
            if choix_mode == "test":
                # On copie le rapport test dans le dossier Calisto

                shutil.copy(Path(Path(__file__).parent.absolute(), "mode_test", "fichier_test", "Report.pdf"),
                            Path(chemin_calisto, "Report.pdf"))
                file_oldname = Path(chemin_calisto, "Report.pdf")
                file_newname_newfile = Path(chemin_calisto, nom_maison_retraite,
                                            nom_patient.upper() + "-" + prenom_patient.title() + "-" + nom_maison_retraite.upper() + ".pdf")
                shutil.move(file_oldname, file_newname_newfile)
                # on met à jour le fichier excel avec identité patient et accompagnant
                ws_liste_referent.cell(row=numero_patient + 4, column=2).value = nom_patient
                ws_liste_referent.cell(row=numero_patient + 4, column=3).value = prenom_patient
                ws_liste_referent.cell(row=numero_patient + 4, column=5).value = nom_accompagnant
                ws_liste_referent.cell(row=numero_patient + 4, column=6).value = prenom_accompagnant
                ws_liste_referent.cell(row=numero_patient + 4, column=7).value = telephone_accompagnant
                ws_liste_referent.cell(row=numero_patient + 4, column=8).value = mail_accompagnant
                # on indique si il y a des empreintes
                if empreinte_OG < 6 and empreinte_OD < 6:
                    ws_liste_referent.cell(row=numero_patient + 4, column=9).value = "Aucune"
                elif empreinte_OG < 6 and empreinte_OD > 5:
                    ws_liste_referent.cell(row=numero_patient + 4, column=9).value = "OD"
                elif empreinte_OG > 5 and empreinte_OD < 6:
                    ws_liste_referent.cell(row=numero_patient + 4, column=9).value = "OG"
                elif empreinte_OG > 5 and empreinte_OD > 5:
                    ws_liste_referent.cell(row=numero_patient + 4, column=9).value = "ODG"

                # On incrémente les variables
                step_one = 0
                step_two = 0
                numero_patient += 1
            else:
                # on ferme la fiche note
                pyautogui.click(709, 411)
                time.sleep(0.1)
                # On force l'audiométrie tonale pour récupérer les données d'audiométrie
                pyautogui.click(25, 92)
                time.sleep(0.1)
                # On valide la confirmation d'enregistrement dans le cas où on revient sur la fiche
                pyautogui.click(928, 566)
                time.sleep(0.1)
                # on enregistre le rapport en pdf
                pyautogui.click(122, 45)
                time.sleep(0.1)
                # on enregistre la session
                pyautogui.click(148, 45)
                time.sleep(2)
                # On verifie que le rapport est a bien été enregistré
                Path(chemin_calisto, "Report.pdf")
                if os.path.exists(Path(chemin_calisto, "Report.pdf")):
                    # on modifie le nom du pdf generé avec le nom du patient
                    file_oldname = Path(chemin_calisto, "Report.pdf")
                    file_newname_newfile = Path(chemin_calisto, nom_maison_retraite,
                                                nom_patient.upper() + "-" + prenom_patient.title() + "-" + nom_maison_retraite.upper() + ".pdf")
                    shutil.move(file_oldname, file_newname_newfile)
                    # on ferme la session
                    pyautogui.click(1897, 12)
                    time.sleep(0.1)
                    # on remet la souris au milieu haut de l'écran
                    pyautogui.moveTo(1000, 350)
                    # on met à jour la liste des référents avec identité patient et accompagnant
                    ws_liste_referent.cell(row=numero_patient + 4, column=2).value = nom_patient.upper()
                    ws_liste_referent.cell(row=numero_patient + 4, column=3).value = prenom_patient.upper()
                    ws_liste_referent.cell(row=numero_patient + 4, column=5).value = nom_accompagnant.upper()
                    ws_liste_referent.cell(row=numero_patient + 4, column=6).value = prenom_accompagnant.upper()
                    ws_liste_referent.cell(row=numero_patient + 4, column=7).value = telephone_accompagnant
                    ws_liste_referent.cell(row=numero_patient + 4, column=8).value = mail_accompagnant.lower()
                    # on indique si il y a des empreintes

                    empreinte_OG = app3.id_reponse[9]
                    empreinte_OD = app3.id_reponse[10]
                    if empreinte_OG == "":
                        empreinte_OG = 1
                    if empreinte_OD == "":
                        empreinte_OD = 1

                    if empreinte_OG < 6 and empreinte_OD < 6:
                        ws_liste_referent.cell(row=numero_patient + 4, column=9).value = "Aucune"
                    elif empreinte_OG < 6 and empreinte_OD > 5:
                        ws_liste_referent.cell(row=numero_patient + 4, column=9).value = "OD"
                    elif empreinte_OG > 5 and empreinte_OD < 6:
                        ws_liste_referent.cell(row=numero_patient + 4, column=9).value = "OG"
                    elif empreinte_OG > 5 and empreinte_OD > 5:
                        ws_liste_referent.cell(row=numero_patient + 4, column=9).value = "ODG"

                    # on met à jour la synthèse
                    # Nom prénom
                    ws_synthese_depistage.cell(row=numero_patient + 6,
                                               column=2).value = nom_patient.upper() + " " + prenom_patient.upper()
                    ws_synthese_depistage.cell(row=numero_patient + 6, column=2).font = defaut_color
                    # Résultat du dépistage
                    # Description de la perte
                    if app3.text_reponse[0] == "Refuse le dépistage":
                        resultat_test = "Refus du dépistage"
                    elif app3.text_reponse[0] == "Dépistage impossible":
                        resultat_test = "Dépistage impossible"
                    elif app3.text_reponse[7] == "OG non appareillable" \
                            and app3.text_reponse[8] == "OD non appareillable":
                        resultat_test = "Non appareillable ODG"
                    elif app3.text_reponse[7] == "OG non appareillable":
                        resultat_test = "Non appareillable OG / " + analyse_perte[0]
                    elif app3.text_reponse[8] == "OD non appareillable":
                        resultat_test = "Non appareillable OD / " + analyse_perte[1]
                    else:
                        resultat_test = analyse_perte[2]

                    # Description état de l'otoscopie
                    if app3.text_reponse[3] == "Cerumen gênant OG" and app3.text_reponse[4] == "Cerumen gênant OD":
                        resultat_test = resultat_test + " ODG cérumen à retirer"
                    elif app3.text_reponse[3] == "Bouchon OG" and app3.text_reponse[4] == "Bouchon OD":
                        resultat_test = resultat_test + " ODG bouchons cerumen"
                    elif app3.text_reponse[3] == "Cerumen gênant OG":
                        resultat_test = resultat_test + " OG cérumen à retirer"
                    elif app3.text_reponse[4] == "Cerumen gênant OD":
                        resultat_test = resultat_test + " OD cérumen à retirer"
                    elif app3.text_reponse[3] == "Bouchon OG":
                        resultat_test = resultat_test + " OG bouchons cerumen"
                    elif app3.text_reponse[4] == "Bouchon OD":
                        resultat_test = resultat_test + " OD bouchons cerumen"
                    ws_synthese_depistage.cell(row=numero_patient + 6, column=3).value = resultat_test
                    ws_synthese_depistage.cell(row=numero_patient + 6, column=3).font = defaut_color

                    # Déjà appareillé
                    if app3.text_reponse[2] == "":
                        resultat_appareillage = "Pas de réponse patient(e)"
                    elif app3.text_reponse[2] == "NON":
                        resultat_appareillage = "NON"
                    else:
                        resultat_appareillage = "OUI"
                    ws_synthese_depistage.cell(row=numero_patient + 6, column=4).value = resultat_appareillage
                    ws_synthese_depistage.cell(row=numero_patient + 6, column=4).font = defaut_color

                    # Définition du besoin
                    if app3.text_reponse[0] == "Refuse le dépistage" or app3.text_reponse[0] == "Dépistage impossible":
                        ws_synthese_depistage.cell(row=numero_patient + 6, column=5).font = impossible_color
                        ws_synthese_depistage.cell(row=numero_patient + 6, column=4).font = impossible_color
                        ws_synthese_depistage.cell(row=numero_patient + 6, column=3).font = impossible_color
                        ws_synthese_depistage.cell(row=numero_patient + 6, column=2).font = impossible_color

                    elif app3.text_reponse[7] == "OG non appareillable" \
                            and app3.text_reponse[8] == "OD non appareillable":
                        ws_synthese_depistage.cell(row=numero_patient + 6, column=5).font = impossible_color
                        ws_synthese_depistage.cell(row=numero_patient + 6, column=4).font = impossible_color
                        ws_synthese_depistage.cell(row=numero_patient + 6, column=3).font = impossible_color
                        ws_synthese_depistage.cell(row=numero_patient + 6, column=2).font = impossible_color

                    elif analyse_perte[3] == "Oui":
                        ws_synthese_depistage.cell(row=numero_patient + 6, column=5).value = analyse_perte[3]
                        ws_synthese_depistage.cell(row=numero_patient + 6, column=5).font = oui_color
                        ws_synthese_depistage.cell(row=numero_patient + 6, column=4).font = oui_color
                        ws_synthese_depistage.cell(row=numero_patient + 6, column=3).font = oui_color
                        ws_synthese_depistage.cell(row=numero_patient + 6, column=2).font = oui_color

                    elif analyse_perte[3] == "Non" or analyse_perte[3] == "Si besoin ressenti":
                        ws_synthese_depistage.cell(row=numero_patient + 6, column=5).value = analyse_perte[3]
                        ws_synthese_depistage.cell(row=numero_patient + 6, column=5).font = defaut_color
                    # On incrémente les variables
                    # Enregistrement ligne par ligne dans la liste et la synthèse

                    # Test fermeture d'excel en cas d'ouverture manuel par l'utilisateur pour éviter les bugs
                    excel_close_test()

                    # On effectue un enregistrment temporaire en cas de bug et on réouvre la fiche (liste référent)
                    wb_liste_referent.save(liste_enregistre_tmp)

                    # On réouvre
                    wb_liste_referent = openpyxl.load_workbook(liste_enregistre_tmp)
                    ws_liste_referent = wb_liste_referent['Feuil1']
                    sheet_liste_referent = wb_liste_referent.active

                    # On effectue un enregistrement temporaire en cas de bug et on réouvre la fiche (synthèse)
                    wb_synthese_depistage.save(synthese_enregistre_tmp)
                    # On réouvre
                    wb_synthese_depistage = openpyxl.load_workbook(synthese_enregistre_tmp)
                    ws_synthese_depistage = wb_synthese_depistage['Feuil1']
                    sheet_synthese_depistage = wb_synthese_depistage.active
                    print(nom_patient + " " + prenom_patient + " bien enregistré")

                    step_one = 0
                    step_two = 0
                    numero_patient += 1
                else:
                    print("Le fichier à mal été enregistré")
                    showMessage("Le fichier à mal été enregistré retournez à l'étape précédente où ré-enregistrez",
                                type='warning', timeout=1000)
                    # sys.exit()
        else:
            print("Il faut compléter les étapes précédentes")
            showMessage("Il faut compléter les étapes précédentes", type='warning', timeout=1000)
            # sys.exit()

    if key == Key.f7:
        print(pyautogui.position())

    if key == Key.f5:
        print("On reinitialise les etapes")
        showMessage("On reinitialise à l'étape 1")
        pygame.mixer.init()
        pygame.mixer.music.load('files/son_fin.mp3')
        pygame.mixer.music.play()
        time.sleep(0.5)
        step_one = 0
        step_two = 0

    if key == Key.f6:
        print("On reinitialise l'étape 2")
        showMessage("On reinitialise l'étape 2")
        pygame.mixer.init()
        pygame.mixer.music.load('files/son_fin.mp3')
        pygame.mixer.music.play()
        time.sleep(0.5)
        pygame.mixer.music.play()
        time.sleep(0.5)
        step_two = 0


# Fermeture avec la touche esc
def press_off(key):
    global chemin_calisto, nom_maison_retraite, dossier_sauvegarde, la_date_jour_save, wb_liste_referent, \
        wb_synthese_depistage, lafin
    if key == Key.esc:
        if len(os.listdir(Path(chemin_calisto, nom_maison_retraite))) == 0:
            print("Le répertoire est vide")
            showMessage("Le répertoire est vide, il n'y a rien à enregistrer, fermeture du programme")
            lafin = 1
            sys.exit()
        else:
            print("fermeture du programme")
            pygame.mixer.init()
            pygame.mixer.music.load('files/son_fin.mp3')
            pygame.mixer.music.play()
            time.sleep(0.5)
            pygame.mixer.music.play()
            time.sleep(0.5)
            pygame.mixer.music.play()
            time.sleep(0.5)
            pygame.mixer.quit()
            # Création fichier pdf fusionné de tous les patients
            print("Création du pdf fusionné de tous les patients")
            try:
                pdf_merge(rf"{chemin_calisto}/{nom_maison_retraite}",
                          rf"{chemin_calisto}/{nom_maison_retraite}/Fiches_tous_patients_{nom_maison_retraite}.pdf")
            except Exception as e:
                print("pdferreur", e)

            lafin = 0
            listener.stop()
            return 0

    if key == Key.f4:
        showMessage("Fermeture de la session avec enregistrement temporaire")
        lafin = 1
        listener.stop()
        return 0


# On reste en veille des touches du clavier
with Listener(on_press=press_on, on_release=press_off) as listener:
    listener.join()

if lafin == 1:
    print("Programme terminé")
    sys.exit()
else:

    # Fin du programme enregistrement
    # On enregistre les fichiers excel Liste et Synthese dans le dossier de la maison de retraite
    print("Enregistrement des fichiers excel List et Synthèse")

    # Liste
    liste_enregistre = Path(chemin_calisto, nom_maison_retraite,
                            "ListeRef-" + la_date_jour_save + "-" + nom_maison_retraite + ".xlsx")

    # Synthese
    synthese_enregistre = Path(chemin_calisto, nom_maison_retraite,
                               "Synthese-" + la_date_jour_save + "-" + nom_maison_retraite + ".xlsx")

    # Test fermeture d'excel en cas d'ouverture manuel par l'utilisateur pour éviter les bugs
    excel_close_test()

    wb_synthese_depistage.save(synthese_enregistre_tmp)
    wb_liste_referent.save(liste_enregistre_tmp)

    excel_triage(liste_enregistre_tmp, synthese_enregistre_tmp)

    # On renomme les fichiers temporaire en fichier definitif
    print('Supression des fichiers excel temporaires')
    shutil.move(liste_enregistre_tmp, liste_enregistre)
    shutil.move(synthese_enregistre_tmp, synthese_enregistre)

    # On renomme le dossier de la maison de retraite avec date jour zone
    print('Mise à la date du dossier de la maison de retraite')
    nom_dossier_sauvegarde = nom_maison_retraite + "-" + datetime.today().strftime('%d-%m-%Y--%H-%M-%S')
    dossier_maison_retraite = Path(chemin_calisto, nom_maison_retraite)
    dossier_maison_retraite_date = Path(chemin_calisto, nom_dossier_sauvegarde)
    dossier_sauvegarde_date = Path(dossier_sauvegarde, nom_dossier_sauvegarde)
    shutil.move(dossier_maison_retraite, dossier_maison_retraite_date)
    # On genere un fichier zip contenant les CR + l'excel
    print("Génération sauvegarde .zip")
    shutil.make_archive(dossier_maison_retraite_date, 'zip', chemin_calisto, nom_dossier_sauvegarde)
    # On déplace l'archive dans le dossier
    print("Sauvegarde copie")
    chemin_archive = Path(chemin_calisto, nom_dossier_sauvegarde + ".zip")
    print(chemin_archive)
    shutil.move(chemin_archive, dossier_maison_retraite_date)
    # On déplace le dossier avec les comptes rendus + fichier excel + version ziper recapitulatif dans le dossier de sauvegarde
    shutil.move(dossier_maison_retraite_date, dossier_sauvegarde_date)
    showMessage("Session terminée, comptes rendu enregistrés")

    print("Programme terminé")
    sys.exit()
