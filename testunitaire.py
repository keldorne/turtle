"""
monmot = "sAgr AnDmeRe"
monmot = monmot.capitalize()
print(monmot.capitalize())
print(monmot.upper())
print(monmot.title())
"""
import textwrap

"""
pygame.mixer.init()
pygame.mixer.music.load('files/son_fin.mp3')
pygame.mixer.music.play()
time.sleep(0.5)
pygame.mixer.music.play()
time.sleep(0.5)
"""

"""
import pickle
import os

names = ["olaf"]
ulf = ["un vrai", "beaugoss"]
names.extend(ulf)
print(names)
pickle.dump(names, open("test.dat", "wb"))
saves = pickle.load(open("test.dat", "rb"))

print(saves)
"""

"""
from fonctiondekev import loss_noah_extractor
import time
import pyautogui
time.sleep(4)
pyautogui.press('f2')
"""

"""
import cv2
import pytesseract
import re
import pyscreenshot
import math
import numpy as np
import time


# img = cv2.imread('imageaudiovoid.png')

# Cette fonction prend une capture d'écran d'une audiométrie sur Calisto et extraie la perte d'audition en dB
def loss_noah_extractor():
    # Capture de l'image
    image = np.array(pyscreenshot.grab().convert('RGB'))
    try:
        print(image.shape)
    except:
        pass

    #image = cv2.imread('imageaudiovoid.png')

    # Traitement de l'image
    # Recadrage, binarisation de l'oreille droite
    imageOD = image[260:285, 225:425]
    imageOD = cv2.cvtColor(imageOD, cv2.COLOR_BGR2GRAY)
    imageOD = cv2.threshold(imageOD, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)[1]

    # Recadrage, binarisation de l'oreille gauche
    imageOG = image[260:285, 1750:1925]
    imageOG = cv2.cvtColor(imageOG, cv2.COLOR_BGR2GRAY)
    imageOG = cv2.threshold(imageOG, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)[1]

    # Extraction des textes
    textOD = pytesseract.image_to_string(imageOD)
    textOG = pytesseract.image_to_string(imageOG)

    #Extraction des valeurs des pertes
    perteod = re.findall("[0-9,-]+", textOD)
    perteog = re.findall("[0-9,-]+", textOG)

    #Mise en forme des valeurs
    if perteod[0] == '-':
        pod = 0
        perteod = 1000
    else:
        pod = 1
        perteod = math.ceil(float(perteod[0].replace(",", ".")))

    if perteog[0] == '-':
        pog = 0
        perteog = 2000
    else:
        pog = 1
        perteog = math.ceil(float(perteog[0].replace(",", ".")))

    #Logique d'analyse de la perte et retour du texte rapport
    if perteod == perteog:
        moyenne_perte = (perteod + perteog) / 2
    elif perteod > perteog:
        moyenne_perte = math.ceil((3*perteod + 7*perteog) / 10)
    elif perteod < perteog:
        moyenne_perte = math.ceil((7*perteod + 3*perteog) / 10)

    if 200 < abs(perteod-perteog):
        analyse_symetrie = ""
        crit_sym = 4
    if 19 < abs(perteod-perteog) < 120:
        analyse_symetrie = "asymétrique"
        crit_sym = 3
    elif abs(perteod-perteog) < 6:
        analyse_symetrie = "symétrique"
        crit_sym = 1
    elif 5 < abs(perteod-perteog) < 20:
        analyse_symetrie = "bilatérale"
        crit_sym = 2
    oreilles = [perteod, perteog, moyenne_perte]
    oreilles_reponse = ["", "", ""]
    degre_reponse = ["", "", ""]
    for count, oreille in enumerate(oreilles):
        if count == 0:
            text_choix_oreille = 'OD'
        elif count == 1:
            text_choix_oreille = 'OG'
        elif count == 2:
            text_choix_oreille = 'ODG'

        if oreille == 1000 or oreille == 2000:
            oreilles_reponse[count] = f"{text_choix_oreille} courbe non caractérisable"
            degre_reponse[count] = 8
        elif oreille < 20:
            oreilles_reponse[count] = f"{text_choix_oreille} audition normale"
            degre_reponse[count] = 0
        elif 20 < oreille < 31:
            oreilles_reponse[count] = f"{text_choix_oreille} perte légère de {oreilles[count]} dB"
            degre_reponse[count] = 1
        elif 30 < oreille < 41:
            oreilles_reponse[count] = f"{text_choix_oreille} perte légère-moyenne de {oreilles[count]} dB"
            degre_reponse[count] = 2
        elif 40 < oreille < 56:
            oreilles_reponse[count] = f"{text_choix_oreille} perte moyenne de {oreilles[count]} dB"
            degre_reponse[count] = 3
        elif 55 < oreille < 71:
            oreilles_reponse[count] = f"{text_choix_oreille} perte moyenne-sévère de {oreilles[count]} dB"
            degre_reponse[count] = 4
        elif 70 < oreille < 81:
            oreilles_reponse[count] = f"{text_choix_oreille} perte sévère de {oreilles[count]} dB"
            degre_reponse[count] = 5
        elif 80 < oreille < 91:
            oreilles_reponse[count] = f"{text_choix_oreille} perte sévère-profonde de {oreilles[count]} dB"
            degre_reponse[count] = 6
        elif 90 < oreille < 119:
            oreilles_reponse[count] = f"{text_choix_oreille} perte profonde de {oreilles[count]} dB"
            degre_reponse[count] = 7

    if perteod > 20 and perteog > 20 and crit_sym == 4:
        analyse = oreilles_reponse[0] + " / " + oreilles_reponse[1]

    elif perteod < 21 and perteog < 21:
        analyse = oreilles_reponse[2]

    elif (perteod < 21 and 20 < perteog) or (20 < perteod and perteog < 21) and (crit_sym == 2 or crit_sym == 3):
        analyse = "Surdité unilérale " + oreilles_reponse[0] + " / " + oreilles_reponse[1]

    elif perteod > 20 and perteog > 20 and crit_sym == 1:
        analyse = f"Surdité {analyse_symetrie} " + oreilles_reponse[2]

    elif perteod > 20 and perteog > 20 and crit_sym == 2 and degre_reponse[0] != degre_reponse[1]:
        analyse = f"Surdité {analyse_symetrie} " + oreilles_reponse[0] + " / " + oreilles_reponse[1]

    elif perteod > 20 and perteog > 20 and crit_sym == 2 and degre_reponse[0] == degre_reponse[1]:
        analyse = f"Surdité {analyse_symetrie} " + oreilles_reponse[2]

    elif perteod > 20 and perteog > 20 and crit_sym == 3:
        analyse = f"Surdité {analyse_symetrie} " + oreilles_reponse[0] + " / " + oreilles_reponse[1]

    print(analyse)
    return analyse
"""
"""
import time
from fonctiondekev import loss_noah_extractor
time.sleep(4)
texts = loss_noah_extractor()
print(texts[0])
print(texts[1])
print(texts[2])
print(texts[3])
"""
"""
from tkinter import *

root = Tk()
t = Text(wrap=WORD, height=3)
t.pack()
root.mainloop()
"""
textedavant="Que c'est bon de dormir dans tes bras, passer la nuit tout contre toi, sentir la douceur de ta peau qui se frotte à la mienne dans notre sommeil. J’aime m'endormir chaque soir dans tes bras après avoir fait l'amour, on est épuisé mais tellement heureux, j'aime qu'après s'être aimé comme on le fait, tu prennes mon visage entre tes mains, tu m'embrasse et tu le déposes contre tes seins, j'aime passer chaque nuit comme ça. Je me sens tellement unique, tellement à toi, je suis ton homme, je suis en ta possession et c'est tellement merveilleux comme sensation.\n"
textdavant = textedavant.strip('\n')
textewrapper = textwrap.wrap(textdavant, width=110)
textedapres = ""
for ligneremarque in textewrapper:
    textedapres = textedapres + ligneremarque + "\n"

print(textedapres)





"""
import openpyxl
from fonctiondekev import excel_triage
excel_triage("montest1.xlsx", "montest2.xlsx")
numero_patient = 0
lechemin = "montest1.xlsx"
wb_liste_referent = openpyxl.load_workbook(lechemin)
ws_liste_referent = wb_liste_referent['Feuil1']
sheet_liste_referent = wb_liste_referent.active

for row in sheet_liste_referent.iter_rows(min_row = 4, max_row = 204, min_col = 2, max_col = 2):
  for cell in row:
    if not cell.value:
        break
    else:
        numero_patient += 1
        print(cell.value)

print (numero_patient)
#print(sheet_liste_referent["A3"].value)
"""

"""
import os

def excel_close_test():
    print("Verification bonne fermeture d'excel")
    try:
      os_cmd = "taskkill /F /IM excel.exe"
      if os.system(os_cmd) != 0:
          raise Exception('Excel est déjà fermé')
      print("Fermeture excel forcée pour enregistrement")
    except:
      print("Excel est déjà bien fermé")
"""
"""
maliste = [3,4,5]
if not maliste:
    print("vide")
else:
    print("plein")

print(os.path.join(chemin_calisto, "ListeRef-" + la_date_jour_save + "-" + nom_maison_retraite + "_KPERREAUT.xlsx"))
lejourseleve = os.path.join(chemin_calisto, "ListeRef-" + la_date_jour_save + "-" + nom_maison_retraite + "_KPERREAUT.xlsx")

liste_enregistre_abs = os.path.abspath(liste_enregistre)
wb_liste_referent = openpyxl.load_workbook(liste_enregistre_abs)
ws_liste_referent = wb_liste_referent['Feuil1']
sheet_liste_referent = wb_liste_referent.active
sheet_liste_referent.auto_filter.ref = "B3:I3"
wb_liste_referent.save(liste_enregistre_abs)



excel = win32com.client.Dispatch("Excel.Application")


#liste_enregistre_abs = Path(liste_enregistre_temporaire).resolve()

print(liste_enregistre_abs)
print(type(liste_enregistre_abs))
wb = excel.Workbooks.Open(liste_enregistre_abs)
ws = wb.Worksheets('Feuil1')
ws.Range('B4:I28').Sort(Key1=ws.Range('B3'), Order1=1, Orientation=1)
wb.Save()
excel.Application.Quit()
time.sleep(0.1)
#shutil.move(liste_enregistre_temporaire, liste_enregistre)


"""
#from fonctiondekev import pdf_merge
#pdf_merge(r"C:\Users\Audio69\Documents\DocOdipro\SynologyDrive\Depistages\MR TESTPDF-17-03-2022--22-36-57", r"C:\Users\Audio69\Documents\DocOdipro\SynologyDrive\Depistages\MR TESTPDF-17-03-2022--22-36-57\touspatient.pdf")


import openpyxl
import win32com.client
import time
from os.path import abspath
from pathlib import Path
import os
import shutil
from fonctiondekev import excel_triage
"""
input()
chemin_calisto = Path(Path.home(), "Documents", "DocumentCalisto")
nom_maison_retraite = "MR VILLA SAINT AGNES"
la_date_jour_save = "31-03-2022"

liste_enregistre = Path(chemin_calisto, nom_maison_retraite,
                                    "ListeRef-" + la_date_jour_save + "-" + nom_maison_retraite + "_KPERREAUT.xlsx")

synthese_enregistre = Path(chemin_calisto, nom_maison_retraite,
                           "Synthese-" + la_date_jour_save + "-" + nom_maison_retraite + "_KPERREAUT.xlsx")

excel_triage(liste_enregistre, synthese_enregistre)
"""




